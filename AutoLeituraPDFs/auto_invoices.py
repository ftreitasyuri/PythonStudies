import os
from openpyxl import Workbook
import pdfplumber
import re
from datetime import datetime
# Verificar quantos arquivos/pdf existem

diretorio = 'pdf_invoices'
arquivos = os.listdir(diretorio)

qtd_arquivos = len(arquivos)

# Debug
# print(qtd_arquivos)
if qtd_arquivos == 0:
    raise Exception("Nenhum arquivo encontrado na pasta")


# Criando estrutura do arquivo excel


wb = Workbook()
ws = wb.active
ws.title = 'Importacao de Nfs'
# Nomeando as colunas
ws['A1'] = 'Nf #'
ws['B1'] = 'Data'
ws['C1'] = "Nome do Arquivo"
ws['D1'] = 'Status'

ultima_linha_vazia = 1
# Loop para descobrir a última linha vazia na planilha, vamos usar para preencher com os dados das nfs
while ws['A' + str(ultima_linha_vazia)].value is not None:
    ultima_linha_vazia += 1
    
# Loop para percorrer e resgatar o conteúdo de cada pdf
for arquivo in arquivos:
    with pdfplumber.open(diretorio + "/" + arquivo) as pdf:
        primeira_pagina = pdf.pages[0]
        texto_pdf = primeira_pagina.extract_text()
        # Debug
        # print(texto_pdf)
        
        # Regex para pegar apenas o conteúdo desejado
    numero_nota_padra = r'INVOICE #(\d+)'
    numero_data_padra = r'DATE (\d{2}/\d{2/\d{4}})'
    
    match_number = re.search(numero_data_padra, texto_pdf)
    match_date = re.search(numero_data_padra, texto_pdf)
    
    # Verificar se foi encontrado
    if match_number:
        numero_nota = match_number.group(1)
        ws['A' + str(ultima_linha_vazia)] = numero_nota
    else:
        ws['A' + str(ultima_linha_vazia)] = 'Não foi encontrado o número da nota'
        
    if match_date:
        numero_data = match_date.group(1)
        ws['B' + str(ultima_linha_vazia)] = numero_data
    else:
        ws['B' + str(ultima_linha_vazia)] = 'Não foi encontrada a data'
        
    ws['C' + str(ultima_linha_vazia)] = arquivo
    ws['D' + str(ultima_linha_vazia)] = "Arquivo lido com sucesso!"
    
    ultima_linha_vazia += 1
    
# Pegando a data e hora para salvar o excel

data_atual = str(datetime.now()).replace(":", "-")
indice_ponto = data_atual.index(".")
data_real = data_atual[:indice_ponto]
    
    
    
        
        
wb.save(f"Notas - {data_real}.xlsx")




